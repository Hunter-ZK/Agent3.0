import sqlite3

from openpyxl import Workbook

from sql_pilot_engine.metadata.ingestion.excel import (
    import_metadata_excel,
)


def test_excel_import_persists_metadata(
    tmp_path,
):

    excel_path = (
        tmp_path
        / "metadata.xlsx"
    )

    database_path = (
        tmp_path
        / "metadata.db"
    )

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "每个表的字段"

    sheet.append(
        (
            "字段英文",
            "字段中文",
            "英文表名",
            "中文表名",
        )
    )

    sheet.append(
        (
            "loan_bal_rmb",
            "贷款余额",
            "dwd_hd_201_cldwdk",
            "绿色单位贷款明细宽表",
        )
    )

    sheet.append(
        (
            "fin_org_type_code",
            "机构类型代码",
            "dwd_hd_201_cldwdk",
            "绿色单位贷款明细宽表",
        )
    )

    workbook.save(
        excel_path
    )

    result = import_metadata_excel(
        excel_path,

        database_path,

        snapshot_label="2026-05",
    )

    assert result.table_count == 1

    assert result.column_count == 2

    assert result.raw_rows == 2

    with sqlite3.connect(
        database_path
    ) as connection:

        table_row = (
            connection.execute(
                """
                SELECT
                    full_name,
                    description,
                    layer
                FROM metadata_table
                """
            )
            .fetchone()
        )

        assert table_row == (
            "dwd_hd_201_cldwdk",
            "绿色单位贷款明细宽表",
            "dwd",
        )

        column_count = (
            connection.execute(
                """
                SELECT COUNT(*)
                FROM metadata_column
                """
            )
            .fetchone()[0]
        )

        assert column_count == 2


def test_excel_import_removes_exact_duplicates(
    tmp_path,
):

    excel_path = (
        tmp_path
        / "metadata.xlsx"
    )

    database_path = (
        tmp_path
        / "metadata.db"
    )

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "每个表的字段"

    sheet.append(
        (
            "字段英文",
            "字段中文",
            "英文表名",
            "中文表名",
        )
    )

    row = (
        "loan_bal_rmb",
        "贷款余额",
        "dwd_hd_201_cldwdk",
        "绿色单位贷款明细宽表",
    )

    sheet.append(row)
    sheet.append(row)

    workbook.save(
        excel_path
    )

    result = import_metadata_excel(
        excel_path,

        database_path,

        snapshot_label="2026-05",
    )

    assert result.raw_rows == 2

    assert result.accepted_rows == 1

    assert result.duplicate_rows == 1

    assert result.column_count == 1


def test_excel_import_keeps_one_primary_description(
    tmp_path,
):

    excel_path = (
        tmp_path
        / "metadata.xlsx"
    )

    database_path = (
        tmp_path
        / "metadata.db"
    )

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "每个表的字段"

    sheet.append(
        (
            "字段英文",
            "字段中文",
            "英文表名",
            "中文表名",
        )
    )

    # 同一个物理字段出现两个中文说明。
    # 当前规则在并列情况下保留首次出现的描述。
    sheet.append(
        (
            "loan_bal_rmb",
            "贷款余额",
            "dwd_hd_201_cldwdk",
            "绿色单位贷款明细宽表",
        )
    )

    sheet.append(
        (
            "loan_bal_rmb",
            "贷款余额折人民币",
            "dwd_hd_201_cldwdk",
            "绿色单位贷款明细宽表",
        )
    )

    workbook.save(
        excel_path
    )

    result = import_metadata_excel(
        excel_path,

        database_path,

        snapshot_label="2026-05",
    )

    # 一个物理字段只入库一次。
    assert result.column_count == 1

    with sqlite3.connect(
        database_path
    ) as connection:

        description = (
            connection.execute(
                """
                SELECT description
                FROM metadata_column
                WHERE name = ?
                """,
                (
                    "loan_bal_rmb",
                ),
            )
            .fetchone()[0]
        )

    assert (
        description
        == "贷款余额"
    )


def test_new_import_becomes_active_batch(
    tmp_path,
):

    excel_path = (
        tmp_path
        / "metadata.xlsx"
    )

    database_path = (
        tmp_path
        / "metadata.db"
    )

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "每个表的字段"

    sheet.append(
        (
            "字段英文",
            "字段中文",
            "英文表名",
            "中文表名",
        )
    )

    sheet.append(
        (
            "id",
            "编号",
            "dwd_test",
            "测试表",
        )
    )

    workbook.save(
        excel_path
    )

    first = import_metadata_excel(
        excel_path,

        database_path,

        snapshot_label="2026-05",
    )

    second = import_metadata_excel(
        excel_path,

        database_path,

        snapshot_label="2026-06",
    )

    with sqlite3.connect(
        database_path
    ) as connection:

        active_batch_id = (
            connection.execute(
                """
                SELECT id
                FROM metadata_batch
                WHERE is_active = 1
                """
            )
            .fetchone()[0]
        )

    assert (
        active_batch_id
        == second.batch_id
    )

    assert (
        first.batch_id
        != second.batch_id
    )