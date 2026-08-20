from __future__ import annotations

import sqlite3

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(
    frozen=True,
    slots=True,
)
class StandardsImportResult:
    rule_count: int

    canonical_root_count: int


def _clean(
    value,
) -> str:

    if value is None:
        return ""

    return str(
        value
    ).strip()


def _sheet_rows(
    workbook,
    sheet_name: str,
):
    if (
        sheet_name
        not in workbook.sheetnames
    ):
        return None, ()

    sheet = workbook[
        sheet_name
    ]

    rows = sheet.iter_rows(
        values_only=True
    )

    headers = tuple(
        _clean(value)
        for value in next(rows)
    )

    indexes = {
        name: index
        for index, name
        in enumerate(headers)
        if name
    }

    return indexes, rows


def _value(
    row,
    indexes: dict[str, int],
    name: str,
) -> str:

    index = indexes.get(
        name
    )

    if index is None:
        return ""

    if index >= len(row):
        return ""

    return _clean(
        row[index]
    )


def import_standards_excel(
    source_path: str | Path,
    database_path: str | Path,
) -> StandardsImportResult:
    """
    将已确认 Standards 导入
    当前正在构建的 metadata.db。

    不导入：
    - 历史词根候选
    - 新增Token候选
    - Review Pool
    - 频率主值

    因为它们不是 Canonical Truth。
    """

    source = Path(
        source_path
    )

    database = Path(
        database_path
    )

    workbook = load_workbook(
        source,
        read_only=True,
        data_only=True,
    )

    rule_count = 0
    root_count = 0

    try:
        with sqlite3.connect(
            database
        ) as connection:

            connection.execute(
                "PRAGMA foreign_keys = ON"
            )

            # ==================================================
            # Governance Rules
            # ==================================================

            indexes, rows = _sheet_rows(
                workbook,
                "治理规则",
            )

            if indexes is not None:

                for row in rows:

                    status = _value(
                        row,
                        indexes,
                        "状态",
                    ).upper()

                    if (
                        status
                        != "CONFIRMED"
                    ):
                        continue

                    rule_code = _value(
                        row,
                        indexes,
                        "编号",
                    )

                    rule_text = _value(
                        row,
                        indexes,
                        "规则",
                    )

                    if (
                        not rule_code
                        or not rule_text
                    ):
                        continue

                    connection.execute(
                        """
                        INSERT INTO standard_rule (
                            rule_code,
                            rule_type,
                            category,
                            rule_text,
                            status,
                            evidence,
                            example,
                            note,
                            source_sheet
                        )
                        VALUES (
                            ?,
                            'governance',
                            ?,
                            ?,
                            ?,
                            '',
                            '',
                            '',
                            '治理规则'
                        )
                        """,
                        (
                            rule_code,
                            _value(
                                row,
                                indexes,
                                "主题",
                            ),
                            rule_text,
                            status,
                        ),
                    )

                    rule_count += 1

            # ==================================================
            # Naming Grammar
            # ==================================================

            indexes, rows = _sheet_rows(
                workbook,
                "命名Grammar",
            )

            if indexes is not None:

                for row in rows:

                    status = _value(
                        row,
                        indexes,
                        "状态",
                    ).upper()

                    if status not in {
                        "CANONICAL",
                        "POLICY",
                    }:
                        continue

                    rule_code = _value(
                        row,
                        indexes,
                        "规则ID",
                    )

                    rule_text = _value(
                        row,
                        indexes,
                        "规则",
                    )

                    if (
                        not rule_code
                        or not rule_text
                    ):
                        continue

                    connection.execute(
                        """
                        INSERT INTO standard_rule (
                            rule_code,
                            rule_type,
                            category,
                            rule_text,
                            status,
                            evidence,
                            example,
                            note,
                            source_sheet
                        )
                        VALUES (
                            ?,
                            'grammar',
                            ?,
                            ?,
                            ?,
                            ?,
                            ?,
                            ?,
                            '命名Grammar'
                        )
                        """,
                        (
                            rule_code,

                            _value(
                                row,
                                indexes,
                                "类别",
                            ),

                            rule_text,

                            status,

                            _value(
                                row,
                                indexes,
                                "数据证据",
                            ),

                            _value(
                                row,
                                indexes,
                                "示例",
                            ),

                            _value(
                                row,
                                indexes,
                                "备注",
                            ),
                        ),
                    )

                    rule_count += 1

            # ==================================================
            # Canonical Root Registry
            #
            # 当前资产若没有该 Sheet：
            # 合法结果就是 0 条。
            #
            # 绝不从历史词根候选推导。
            # ==================================================

            indexes, rows = _sheet_rows(
                workbook,
                "Canonical词根注册表",
            )

            if indexes is not None:

                for row in rows:

                    status = _value(
                        row,
                        indexes,
                        "Status",
                    ).upper()

                    if status not in {
                        "CONFIRMED",
                        "DEPRECATED",
                    }:
                        continue

                    concept = _value(
                        row,
                        indexes,
                        "Canonical Concept",
                    )

                    expression = _value(
                        row,
                        indexes,
                        "中文表达",
                    )

                    root = _value(
                        row,
                        indexes,
                        "Canonical Root",
                    )

                    if (
                        not concept
                        or not expression
                        or not root
                    ):
                        continue

                    connection.execute(
                        """
                        INSERT INTO canonical_root (
                            canonical_concept,
                            chinese_expression,
                            canonical_root,
                            root_type,
                            status,
                            source,
                            note
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            concept,
                            expression,
                            root,

                            _value(
                                row,
                                indexes,
                                "Root Type",
                            ),

                            status,

                            _value(
                                row,
                                indexes,
                                "Source",
                            ),

                            _value(
                                row,
                                indexes,
                                "Note",
                            ),
                        ),
                    )

                    root_count += 1

    finally:
        workbook.close()

    return StandardsImportResult(
        rule_count=rule_count,

        canonical_root_count=(
            root_count
        ),
    )