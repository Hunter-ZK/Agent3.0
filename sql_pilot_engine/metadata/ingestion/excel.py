from __future__ import annotations

import sqlite3

from collections import Counter
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook



REQUIRED_COLUMNS = {
    "字段英文",
    "字段中文",
    "英文表名",
    "中文表名",
}


KNOWN_LAYERS = {
    "ods",
    "dim",
    "dwd",
    "dws",
    "ads",
    "ver",
}


@dataclass(
    frozen=True,
    slots=True,
)
class ExcelMetadataImportResult:
    table_count: int
    column_count: int

    raw_rows: int
    accepted_rows: int
    duplicate_rows: int
    skipped_rows: int


def _clean(
    value,
) -> str:

    if value is None:
        return ""

    return str(value).strip()


def _infer_layer(
    table_name: str,
) -> str:

    prefix = (
        table_name
        .strip()
        .lower()
        .split("_", 1)[0]
    )

    if prefix in KNOWN_LAYERS:
        return prefix

    return ""


def _primary_description(
    counter: Counter[str],
) -> str:
    """
    一个物理对象只保留一个主描述。

    规则：
    1. 出现次数最多；
    2. 次数相同时使用最早出现的描述。

    Counter保持插入顺序，
    因此most_common在并列时会保留首次顺序。
    """

    if not counter:
        return ""

    return counter.most_common(1)[0][0]


def import_metadata_excel(
    source_path: str | Path,
    database_path: str | Path,
) -> ExcelMetadataImportResult:
    """
    将Excel元数据持久化导入SQLite。

    注意：
    这是维护流程，不应该由Agent Runtime调用。
    """

    source = Path(source_path)
    database = Path(database_path)


    workbook = load_workbook(
        source,
        read_only=True,
        data_only=True,
    )

    try:

        sheet = (
            workbook["每个表的字段"]
            if "每个表的字段"
            in workbook.sheetnames
            else workbook.active
        )

        rows = sheet.iter_rows(
            values_only=True
        )

        headers = tuple(
            _clean(value)
            for value in next(rows)
        )

        column_indexes = {
            name: index
            for index, name
            in enumerate(headers)
        }

        missing_columns = (
            REQUIRED_COLUMNS
            - set(column_indexes)
        )

        if missing_columns:
            raise ValueError(
                "Missing required columns: "
                + ", ".join(
                    sorted(missing_columns)
                )
            )

        # --------------------------------------------------
        # 先在内存中按物理表 / 字段聚合。
        #
        # 这里不是Runtime缓存，
        # 只是一次Excel导入过程中的临时结构。
        # --------------------------------------------------

        tables: dict[
            str,
            dict,
        ] = {}

        seen_records: set[
            tuple[
                str,
                str,
                str,
                str,
            ]
        ] = set()

        raw_rows = 0
        accepted_rows = 0
        duplicate_rows = 0
        skipped_rows = 0

        for row in rows:

            raw_rows += 1

            column_name = _clean(
                row[
                    column_indexes[
                        "字段英文"
                    ]
                ]
            )

            column_description = _clean(
                row[
                    column_indexes[
                        "字段中文"
                    ]
                ]
            )

            table_name = _clean(
                row[
                    column_indexes[
                        "英文表名"
                    ]
                ]
            )

            table_description = _clean(
                row[
                    column_indexes[
                        "中文表名"
                    ]
                ]
            )

            if (
                not table_name
                or not column_name
            ):
                skipped_rows += 1
                continue

            normalized_table_name = (
                table_name.lower()
            )

            normalized_column_name = (
                column_name.lower()
            )

            record = (
                normalized_table_name,
                normalized_column_name,
                table_description,
                column_description,
            )

            # 完全相同的原始记录只保留一次。
            if record in seen_records:
                duplicate_rows += 1
                continue

            seen_records.add(record)

            accepted_rows += 1

            table_data = tables.setdefault(
                normalized_table_name,
                {
                    "descriptions": Counter(),
                    "columns": {},
                },
            )

            if table_description:
                table_data[
                    "descriptions"
                ][table_description] += 1

            columns = table_data[
                "columns"
            ]

            if (
                normalized_column_name
                not in columns
            ):
                columns[
                    normalized_column_name
                ] = {
                    "descriptions": Counter(),

                    # 按第一次遇到字段的顺序保存。
                    "ordinal_position": (
                        len(columns) + 1
                    ),
                }

            column_data = columns[
                normalized_column_name
            ]

            if column_description:
                column_data[
                    "descriptions"
                ][column_description] += 1

    finally:

        workbook.close()

    # ------------------------------------------------------
    # 持久化
    #
    # 注意：
    # 这里写入的是 Rebuild 阶段创建的临时数据库。
    # Runtime 不调用这个函数。
    # ------------------------------------------------------

    with sqlite3.connect(
        database
    ) as connection:

        connection.execute(
            "PRAGMA foreign_keys = ON"
        )

        table_count = 0
        column_count = 0

        for (
            table_name,
            table_data,
        ) in tables.items():

            table_cursor = (
                connection.execute(
                    """
                    INSERT INTO metadata_table (
                        full_name,
                        description,
                        layer,
                        row_count,
                        size_bytes
                    )
                    VALUES (
                        ?,
                        ?,
                        ?,
                        NULL,
                        NULL
                    )
                    """,
                    (
                        table_name,

                        _primary_description(
                            table_data[
                                "descriptions"
                            ]
                        ),

                        _infer_layer(
                            table_name
                        ),
                    ),
                )
            )

            table_id = int(
                table_cursor.lastrowid
            )

            table_count += 1

            for (
                column_name,
                column_data,
            ) in table_data[
                "columns"
            ].items():

                connection.execute(
                    """
                    INSERT INTO metadata_column (
                        table_id,
                        name,
                        description,
                        data_type,
                        nullable,
                        ordinal_position,
                        is_partition,
                        distinct_count
                    )
                    VALUES (
                        ?,
                        ?,
                        ?,
                        '',
                        NULL,
                        ?,
                        NULL,
                        NULL
                    )
                    """,
                    (
                        table_id,

                        column_name,

                        _primary_description(
                            column_data[
                                "descriptions"
                            ]
                        ),

                        column_data[
                            "ordinal_position"
                        ],
                    ),
                )

                column_count += 1

    return ExcelMetadataImportResult(
        table_count=table_count,
        column_count=column_count,

        raw_rows=raw_rows,
        accepted_rows=accepted_rows,
        duplicate_rows=duplicate_rows,
        skipped_rows=skipped_rows,
    )